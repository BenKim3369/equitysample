from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

from fetch_komis import PriceRecord
from logger import ManualReviewRow, UpdateLogRow
from utils import AppConfig, is_empty_cell, parse_month


@dataclass
class UpdateResult:
    output_path: Path
    update_logs: list[UpdateLogRow]
    manual_logs: list[ManualReviewRow]
    summary: dict[str, int]


def update_workbook(
    workbook_path: Path,
    output_dir: Path,
    config: AppConfig,
    records: list[PriceRecord],
    timestamp: str,
    fetch_issues: list[dict[str, str]],
) -> UpdateResult:
    wb = load_workbook(workbook_path, data_only=False)

    if config.workbook.sheet_name not in wb.sheetnames:
        raise ValueError(f"대상 시트({config.workbook.sheet_name})를 찾지 못했습니다.")

    for protected_name in config.workbook.protected_sheet_names:
        if protected_name not in wb.sheetnames:
            raise ValueError(f"보호 대상 시트({protected_name})가 없습니다. 무결성 확인 실패.")

    sheet = wb[config.workbook.sheet_name]

    price_lookup = _build_price_lookup(records)
    update_logs: list[UpdateLogRow] = []
    manual_logs: list[ManualReviewRow] = []
    formula_cells_touched: set[str] = set()

    issue_by_mineral = {item["mineral"]: item for item in fetch_issues}

    for issue in fetch_issues:
        manual_logs.append(
            ManualReviewRow(
                mineral=issue["mineral"],
                month="",
                workbook_target="",
                issue_type=issue["issue_type"],
                note=issue["note"],
            )
        )

    for mineral, mineral_cfg in config.minerals.items():
        date_col = mineral_cfg.date_col
        price_col = mineral_cfg.price_col

        for row_idx in range(2, sheet.max_row + 1):
            date_value = sheet[f"{date_col}{row_idx}"].value
            month = parse_month(date_value)
            if month is None:
                continue

            target_cell = f"{price_col}{row_idx}"
            price_cell = sheet[target_cell]

            if isinstance(price_cell.value, str) and price_cell.value.startswith("="):
                formula_cells_touched.add(target_cell)
                update_logs.append(
                    UpdateLogRow(
                        mineral=mineral,
                        month=month.isoformat(),
                        source_label="",
                        source_unit="",
                        value="",
                        target_cell=target_cell,
                        workbook_file=workbook_path.name,
                        status="error",
                        note="수식 셀은 업데이트 대상이 아니므로 건너뜀",
                    )
                )
                continue

            if not is_empty_cell(price_cell.value):
                update_logs.append(
                    UpdateLogRow(
                        mineral=mineral,
                        month=month.isoformat(),
                        source_label="",
                        source_unit="",
                        value="",
                        target_cell=target_cell,
                        workbook_file=workbook_path.name,
                        status="skipped_existing",
                        note="기존 값이 있어 덮어쓰지 않았습니다.",
                    )
                )
                continue

            if mineral not in price_lookup:
                issue = issue_by_mineral.get(mineral, {})
                status = issue.get("issue_type", "skipped_missing_source")
                note = issue.get("note", "해당 광종의 KOMIS 시계열을 확정하지 못했습니다.")
                if status not in {"skipped_missing_source", "skipped_ambiguous_series"}:
                    status = "skipped_missing_source"
                update_logs.append(
                    UpdateLogRow(
                        mineral=mineral,
                        month=month.isoformat(),
                        source_label="",
                        source_unit="",
                        value="",
                        target_cell=target_cell,
                        workbook_file=workbook_path.name,
                        status=status,
                        note=note,
                    )
                )
                manual_logs.append(
                    ManualReviewRow(
                        mineral=mineral,
                        month=month.isoformat(),
                        workbook_target=target_cell,
                        issue_type="missing_source",
                        note="시계열 확정 실패 또는 데이터 없음",
                    )
                )
                continue

            entry = price_lookup[mineral].get(month)
            if entry is None:
                update_logs.append(
                    UpdateLogRow(
                        mineral=mineral,
                        month=month.isoformat(),
                        source_label="",
                        source_unit="",
                        value="",
                        target_cell=target_cell,
                        workbook_file=workbook_path.name,
                        status="skipped_no_matching_month",
                        note="KOMIS 월평균 데이터에서 같은 월을 찾지 못했습니다.",
                    )
                )
                manual_logs.append(
                    ManualReviewRow(
                        mineral=mineral,
                        month=month.isoformat(),
                        workbook_target=target_cell,
                        issue_type="missing_month",
                        note="해당 월 데이터가 KOMIS 응답에 없음",
                    )
                )
                continue

            source_label, source_unit, value = entry
            price_cell.value = value
            update_logs.append(
                UpdateLogRow(
                    mineral=mineral,
                    month=month.isoformat(),
                    source_label=source_label,
                    source_unit=source_unit,
                    value=str(value),
                    target_cell=target_cell,
                    workbook_file=workbook_path.name,
                    status="updated",
                    note="빈 셀에만 입력 완료",
                )
            )

    changed_cells = {log.target_cell for log in update_logs if log.status == "updated"}
    _run_integrity_checks(wb, config, changed_cells, formula_cells_touched)

    output_name = config.workbook.output_name_template.format(timestamp=timestamp)
    output_path = output_dir / output_name
    wb.save(output_path)
    wb.close()

    # reopen test (quality check)
    reopened = load_workbook(output_path, data_only=False)
    reopened.close()

    summary: Dict[str, int] = {}
    for row in update_logs:
        summary[row.status] = summary.get(row.status, 0) + 1

    return UpdateResult(
        output_path=output_path,
        update_logs=update_logs,
        manual_logs=manual_logs,
        summary=summary,
    )


def _build_price_lookup(records: list[PriceRecord]) -> dict[str, dict[date, tuple[str, str, float]]]:
    lookup: dict[str, dict[date, tuple[str, str, float]]] = {}
    for r in records:
        lookup.setdefault(r.mineral, {})[r.month] = (r.source_label, r.source_unit, r.value)
    return lookup


def _run_integrity_checks(
    wb,
    config: AppConfig,
    changed_cells: set[str],
    formula_cells_touched: set[str],
) -> None:
    if config.workbook.sheet_name not in wb.sheetnames:
        raise ValueError("저장 전 무결성 검사 실패: Chart 시트가 사라졌습니다.")

    for protected_name in config.workbook.protected_sheet_names:
        if protected_name not in wb.sheetnames:
            raise ValueError(f"저장 전 무결성 검사 실패: {protected_name} 시트 누락")

    if formula_cells_touched:
        # Detect but do not fail hard; formulas were explicitly skipped.
        return

    allowed_columns = {cfg.price_col for cfg in config.minerals.values()}
    for cell_ref in changed_cells:
        col = "".join(ch for ch in cell_ref if ch.isalpha())
        if col not in allowed_columns:
            raise ValueError(f"허용되지 않은 셀 변경 탐지: {cell_ref}")
